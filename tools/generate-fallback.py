#!/usr/bin/env python3
"""
generate-fallback.py
Генерирует fallback-data.js из rus_pharma_field_advanced.xlsx

Зависимости: pip install openpyxl
Запуск:      python tools/generate-fallback.py
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl не установлен. Запустите: pip install openpyxl")
    sys.exit(1)

# Пути
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.path.join(PROJECT_ROOT, "rus_pharma_field_advanced.xlsx")
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "fallback-data.js")


def safe_str(value):
    """Безопасное преобразование значения в строку."""
    if value is None:
        return ""
    s = str(value).strip()
    # Если значение = "пример" — пустая строка
    if s.lower() == "пример":
        return ""
    return s


def safe_int(value, default=0):
    """Безопасное преобразование в int."""
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def extract_questions(wb):
    """Из листа 'Вопросы' — по 1 вопросу на каждую уникальную компетенцию.
    Предпочитаем сложность 1."""
    ws = wb["Вопросы"]
    # Собираем все вопросы по компетенциям
    comp_questions = {}  # competencyId -> [(row_idx, difficulty), ...]
    for row_idx in range(2, ws.max_row + 1):
        comp = ws.cell(row=row_idx, column=2).value
        if not comp:
            continue
        difficulty = safe_int(ws.cell(row=row_idx, column=9).value, 1)
        if comp not in comp_questions:
            comp_questions[comp] = []
        comp_questions[comp].append((row_idx, difficulty))

    # Для каждой компетенции выбираем вопрос с наименьшей сложностью
    questions = []
    for comp, rows in comp_questions.items():
        # Сортируем по сложности, берём первый
        rows.sort(key=lambda x: x[1])
        row_idx = rows[0][0]

        qid = safe_str(ws.cell(row=row_idx, column=1).value)
        question_text = safe_str(ws.cell(row=row_idx, column=3).value)
        answers_raw = safe_str(ws.cell(row=row_idx, column=4).value)
        correct = safe_int(ws.cell(row=row_idx, column=5).value, 0)
        explanation = safe_str(ws.cell(row=row_idx, column=7).value)
        category = safe_str(ws.cell(row=row_idx, column=10).value)
        module = safe_int(ws.cell(row=row_idx, column=11).value, 1)
        difficulty = safe_int(ws.cell(row=row_idx, column=9).value, 1)

        # "Правильный" — 0-based индекс (проверено: Row2 correct=2 = "Внутривенно")
        answers = [a.strip() for a in answers_raw.split(";") if a.strip()]

        questions.append({
            "id": f"q{qid}",
            "category": category,
            "type": "multiple",
            "question": question_text,
            "answers": answers,
            "correct": correct,
            "explanation": explanation,
            "competencyId": comp,
            "module": module,
            "difficulty": difficulty,
        })

    print(f"  Вопросы: {len(questions)} (по 1 на компетенцию)")
    return questions


def extract_drugs(wb):
    """Из листа 'Препараты' — ВСЕ записи."""
    ws = wb["Препараты"]
    drugs = []
    for row_idx in range(2, ws.max_row + 1):
        drug_id = ws.cell(row=row_idx, column=1).value
        if drug_id is None:
            continue
        drugs.append({
            "id": f"d{safe_int(drug_id)}",
            "name": safe_str(ws.cell(row=row_idx, column=2).value),
            "category": safe_str(ws.cell(row=row_idx, column=3).value),
            "inn": safe_str(ws.cell(row=row_idx, column=4).value),
            "form": safe_str(ws.cell(row=row_idx, column=6).value),
            "dosage": safe_str(ws.cell(row=row_idx, column=7).value),
            "indications": safe_str(ws.cell(row=row_idx, column=8).value),
            "contraindications": safe_str(ws.cell(row=row_idx, column=9).value),
            "sideEffects": safe_str(ws.cell(row=row_idx, column=10).value),
            "fieldNotes": safe_str(ws.cell(row=row_idx, column=11).value),
        })

    print(f"  Препараты: {len(drugs)}")
    return drugs


def extract_scenarios(wb):
    """Из листа 'Сценарии' — CASE_001: START + ключевые 6-8 узлов."""
    ws = wb["Сценарии"]

    # Ключевые узлы для fallback (START + основной путь + win + fail)
    KEY_NODES = {
        "START", "NODE_A", "NODE_B", "NODE_D",
        "NODE_F", "NODE_L", "WIN_1", "WIN_2", "FAIL_1"
    }

    nodes = {}
    for row_idx in range(2, ws.max_row + 1):
        sid = ws.cell(row=row_idx, column=1).value
        if sid != "CASE_001":
            if sid is not None and sid != "CASE_001":
                break
            continue

        node_id = safe_str(ws.cell(row=row_idx, column=2).value)
        if node_id not in KEY_NODES:
            continue

        node_type = safe_str(ws.cell(row=row_idx, column=3).value)
        title = safe_str(ws.cell(row=row_idx, column=4).value)
        description = safe_str(ws.cell(row=row_idx, column=5).value)
        patient_state = safe_str(ws.cell(row=row_idx, column=7).value)

        # Варианты выбора (до 3)
        choices = []
        for i in range(3):
            col_text = 11 + i * 3       # колонки 11, 14, 17
            col_next = 12 + i * 3       # колонки 12, 15, 18
            col_effect = 13 + i * 3     # колонки 13, 16, 19

            text = safe_str(ws.cell(row=row_idx, column=col_text).value)
            next_node = safe_str(ws.cell(row=row_idx, column=col_next).value)
            effect = safe_str(ws.cell(row=row_idx, column=col_effect).value)

            if text and next_node:
                choice = {"text": text, "nextNode": next_node}
                if effect:
                    choice["effect"] = effect
                choices.append(choice)

        nodes[node_id] = {
            "id": node_id,
            "title": title,
            "description": description,
            "type": node_type,
            "patientState": patient_state,
            "choices": choices,
        }

    scenario = {
        "id": "CASE_001",
        "title": "Огнестрельное ранение",
        "nodes": nodes,
    }

    print(f"  Сценарии: 1 (CASE_001, {len(nodes)} узлов)")
    return [scenario]


def generate_js(questions, drugs, scenarios):
    """Генерирует fallback-data.js."""

    def to_js(obj, indent=0):
        """JSON с отступами, совместимый с JS."""
        return json.dumps(obj, ensure_ascii=False, indent=4)

    js = f"""// fallback-data.js
// Сгенерировано автоматически из rus_pharma_field_advanced.xlsx
// Запуск: python tools/generate-fallback.py

const FALLBACK_QUESTIONS = {to_js(questions)};

const FALLBACK_DRUGS = {to_js(drugs)};

const FALLBACK_SCENARIOS = {to_js(scenarios)};
"""
    return js


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: Файл не найден: {XLSX_PATH}")
        sys.exit(1)

    print(f"Читаю: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print("Извлечение данных:")
    questions = extract_questions(wb)
    drugs = extract_drugs(wb)
    scenarios = extract_scenarios(wb)

    js_content = generate_js(questions, drugs, scenarios)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"\nЗаписано: {OUTPUT_PATH}")
    print(f"Размер: {len(js_content)} байт")


if __name__ == "__main__":
    main()
